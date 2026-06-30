from typing import Any, Callable, Dict, List, Set

import pandas as pd
from openpyxl.workbook.workbook import Workbook


class ProcesadorHojaConsultas:
    def __init__(
            self,
            get_target_headers: Callable[[Any, int], Dict[str, int]],
            find_target_col: Callable[[Dict[str, int], str], int | None],
            resolve_source_column: Callable[[List[str], str], str | None],
            to_clean_text: Callable[[Any], str],
            normalize_fecha_yyyy_mm_dd_hh_mm: Callable[[Any], str],
            normalize_two_digit_code: Callable[[Any, Set[str]], str],
            normalize_cod_servicio: Callable[[Any], str],
            normalize_tipo_documento: Callable[[Any], str],
            normalize_amount: Callable[[Any, str], str],
            validate_consulta_row: Callable[[Dict[str, Any]], None],
            allowed_modalidad: Set[str],
            allowed_grupo_servicios: Set[str],
            allowed_finalidad: Set[str],
            allowed_causa: Set[str],
            allowed_tipo_diag_principal: Set[str],
            allowed_concepto_recaudo: Set[str],
    ):
        self._get_target_headers = get_target_headers
        self._find_target_col = find_target_col
        self._resolve_source_column = resolve_source_column
        self._to_clean_text = to_clean_text
        self._normalize_fecha_yyyy_mm_dd_hh_mm = normalize_fecha_yyyy_mm_dd_hh_mm
        self._normalize_two_digit_code = normalize_two_digit_code
        self._normalize_cod_servicio = normalize_cod_servicio
        self._normalize_tipo_documento = normalize_tipo_documento
        self._normalize_amount = normalize_amount
        self._validate_consulta_row = validate_consulta_row

        self._allowed_modalidad = allowed_modalidad
        self._allowed_grupo_servicios = allowed_grupo_servicios
        self._allowed_finalidad = allowed_finalidad
        self._allowed_causa = allowed_causa
        self._allowed_tipo_diag_principal = allowed_tipo_diag_principal
        self._allowed_concepto_recaudo = allowed_concepto_recaudo

    def ejecutar(
            self,
            df_source: pd.DataFrame,
            wb: Workbook,
            config: Dict[str, Any],
            num_documento_id_obligado: str,
            usuarios_ids: Set[str],
    ) -> None:
        ws = wb[str(config["consultas_sheet"])]
        headers = self._get_target_headers(ws, int(config["consultas_header_row"]))

        required_cols = [
            "num_DocumentoIdObligado",
            "consecutivoUsuario",
            "codPrestador",
            "fechaInicioAtencion",
            "numAutorizacion",
            "codConsulta",
            "modalidadGrupoServicioTecSal",
            "grupoServicios",
            "codServicio",
            "finalidadTecnologiaSalud",
            "causaMotivoAtencion",
            "codDiagnosticoPrincipal",
            "codDiagnosticoRelacionado1",
            "codDiagnosticoRelacionado2",
            "codDiagnosticoRelacionado3",
            "tipoDiagnosticoPrincipal",
            "tipoDocumentoIdentificacion",
            "numDocumentoIdentificacion",
            "vrServicio",
            "conceptoRecaudo",
            "valorPagoModerador",
            "numFEVPagoModerador",
            "consecutivo",
        ]

        target_cols: Dict[str, int] = {}
        for col_name in required_cols:
            col_idx = self._find_target_col(headers, col_name)
            if col_idx is None:
                raise ValueError(f"No se encontro la columna destino '{col_name}' en hoja consultas.")
            target_cols[col_name] = col_idx

        source_mapping = {
            "consecutivoUsuario": "Número identificación",
            "codPrestador": "Código Prestador",
            "fechaInicioAtencion": "Fecha y hora",
            "numAutorizacion": "Autorización",
            "codConsulta": "Código",
            "modalidadGrupoServicioTecSal": "Modalidad tecnología salud",
            "grupoServicios": "Grupo servicios",
            "codServicio": "Servicio",
            "finalidadTecnologiaSalud": "Finalidad tecnología",
            "causaMotivoAtencion": "Causa externa",
            "codDiagnosticoPrincipal": "Código Diagnóstico principal",
            "tipoDiagnosticoPrincipal": "Tipo diagnóstico principal",
            "tipoDocumentoIdentificacion": "Tipo ID profesional",
            "numDocumentoIdentificacion": "Número ID profesional",
            "vrServicio": "Valor servicio",
            "conceptoRecaudo": "Concepto recaudo",
            "valorPagoModerador": "Valor pago moderador",
        }

        resolved: Dict[str, str] = {}
        for target_field, source_name in source_mapping.items():
            src_col = self._resolve_source_column(df_source.columns.tolist(), source_name)
            if not src_col:
                raise ValueError(f"No se pudo mapear columna origen '{source_name}' para '{target_field}'.")
            resolved[target_field] = src_col

        errors: List[str] = []
        row_out = int(config["consultas_start_row"])
        consecutivo_por_usuario: Dict[str, int] = {}

        print("Columna usada para consecutivoUsuario:", resolved["consecutivoUsuario"])
        print("Primeros 10 valores:", df_source[resolved["consecutivoUsuario"]].head(10).tolist())

        for row_num_excel, (_, src_row) in enumerate(df_source.iterrows(), start=2):
            consecutivo_usuario = self._to_clean_text(src_row[resolved["consecutivoUsuario"]])

            out = {
                "num_DocumentoIdObligado": self._to_clean_text(num_documento_id_obligado),
                "consecutivoUsuario": consecutivo_usuario,
                "codPrestador": self._to_clean_text(src_row[resolved["codPrestador"]]),
                "fechaInicioAtencion": self._normalize_fecha_yyyy_mm_dd_hh_mm(src_row[resolved["fechaInicioAtencion"]]),
                "numAutorizacion": self._to_clean_text(src_row[resolved["numAutorizacion"]]),
                "codConsulta": self._to_clean_text(src_row[resolved["codConsulta"]]),
                "modalidadGrupoServicioTecSal": self._normalize_two_digit_code(
                    src_row[resolved["modalidadGrupoServicioTecSal"]], self._allowed_modalidad
                ),
                "grupoServicios": self._normalize_two_digit_code(
                    src_row[resolved["grupoServicios"]], self._allowed_grupo_servicios
                ),
                "codServicio": self._normalize_cod_servicio(src_row[resolved["codServicio"]]),
                "finalidadTecnologiaSalud": self._normalize_two_digit_code(
                    src_row[resolved["finalidadTecnologiaSalud"]], self._allowed_finalidad
                ),
                "causaMotivoAtencion": self._normalize_two_digit_code(
                    src_row[resolved["causaMotivoAtencion"]], self._allowed_causa
                ),
                "codDiagnosticoPrincipal": self._to_clean_text(src_row[resolved["codDiagnosticoPrincipal"]]),
                "codDiagnosticoRelacionado1": "",
                "codDiagnosticoRelacionado2": "",
                "codDiagnosticoRelacionado3": "",
                "tipoDiagnosticoPrincipal": self._normalize_two_digit_code(
                    src_row[resolved["tipoDiagnosticoPrincipal"]], self._allowed_tipo_diag_principal
                ),
                "tipoDocumentoIdentificacion": self._normalize_tipo_documento(
                    src_row[resolved["tipoDocumentoIdentificacion"]]
                ),
                "numDocumentoIdentificacion": self._to_clean_text(src_row[resolved["numDocumentoIdentificacion"]]),
                "vrServicio": self._normalize_amount(src_row[resolved["vrServicio"]], default_value="0"),
                "conceptoRecaudo": self._normalize_two_digit_code(
                    src_row[resolved["conceptoRecaudo"]], self._allowed_concepto_recaudo
                ),
                "valorPagoModerador": int(self._normalize_amount(src_row[resolved["valorPagoModerador"]], default_value="0")),
                "numFEVPagoModerador": "",
                "consecutivo": 0,
            }

            if not out["consecutivoUsuario"]:
                errors.append(f"Fila origen {row_num_excel}: consecutivoUsuario vacio.")
                continue

            if out["consecutivoUsuario"] not in usuarios_ids:
                errors.append(
                    f"Fila origen {row_num_excel}: consecutivoUsuario '{out['consecutivoUsuario']}' no existe en hoja usuarios."
                )
                continue

            proximo_consecutivo = consecutivo_por_usuario.get(out["consecutivoUsuario"], 0) + 1
            out["consecutivo"] = proximo_consecutivo

            try:
                self._validate_consulta_row(out)
            except ValueError as e:
                errors.append(f"Fila origen {row_num_excel}: {e}")
                continue

            consecutivo_por_usuario[out["consecutivoUsuario"]] = proximo_consecutivo

            for field in required_cols:
                value_to_write = out.get(field, "")

                if field in {"codServicio", "vrServicio", "valorPagoModerador", "consecutivo"}:
                    if str(value_to_write).strip() != "":
                        value_to_write = int(str(value_to_write))
                    else:
                        value_to_write = None

                cell = ws.cell(row=row_out, column=target_cols[field], value=value_to_write)

                if field in {"codServicio", "vrServicio", "valorPagoModerador", "consecutivo"}:
                    cell.number_format = "0"

            print(
                "row", row_out,
                "user", out["consecutivoUsuario"],
                "calc", out["consecutivo"],
                "cell_before", ws.cell(row=row_out, column=target_cols["consecutivo"]).value
            )
            ws.cell(row=row_out, column=target_cols["consecutivo"], value=out["consecutivo"])
            print("cell_after", ws.cell(row=row_out, column=target_cols["consecutivo"]).value)

            row_out += 1

        if errors:
            msg = "\n".join(errors[:50])
            raise ValueError(
                "Se encontraron filas invalidas en 'consultas'.\n"
                f"Total errores: {len(errors)}\n"
                f"Primeros errores:\n{msg}"
            )
