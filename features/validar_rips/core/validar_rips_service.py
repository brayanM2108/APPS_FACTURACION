import re
import unicodedata
from difflib import get_close_matches
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Set

import pandas as pd
from openpyxl import load_workbook

from features.validar_rips.core.procesar_hoja_transaccion import ProcesadorHojaTransaccion
from features.validar_rips.core.procesar_hoja_usuarios import ProcesadorHojaUsuarios
from features.validar_rips.core.procesar_hoja_consultas import ProcesadorHojaConsultas


# ============================================================
# CONFIG
# ============================================================
CONFIG: Dict[str, Any] = {
    "source_file": r"C:\Users\TECNICOESTADISTICO.P\Downloads\Copia E-370211.xlsx",
    "template_file": r"C:\Users\TECNICOESTADISTICO.P\Downloads\plantilla_fev-rips_v 1_0 (1).xlsm",
    "output_file": r"C:\Users\TECNICOESTADISTICO.P\Downloads\salidaFinalRips.xlsm",

    # Hoja origen de informacion (0 = primera hoja)
    "source_sheet": 0,

    # Hoja 1 destino
    "transaccion_sheet": "transaccion",
    "transaccion_header_row": 1,
    "transaccion_start_row": 2,

    # Hoja 2 destino
    "usuarios_sheet": "usuarios",
    "usuarios_header_row": 1,
    "usuarios_start_row": 2,

    # Hoja 3 destino
    "consultas_sheet": "consultas",
    "consultas_header_row": 1,
    "consultas_start_row": 2,

    # Dato fijo
    "num_documento_id_obligado": "900231829",
}

# ============================================================
# CATALOGOS PERMITIDOS
# ============================================================
ALLOWED_TIPO_NOTA: Set[str] = {"NA", "NC", "ND", "RS"}
ALLOWED_TIPO_DOC: Set[str] = {"AS", "CC", "CD", "CE", "CN", "DE", "MS", "NV", "PA", "PE", "PT", "RC", "SC", "SI", "TI"}
ALLOWED_TIPO_USUARIO: Set[str] = {f"{i:02d}" for i in range(1, 14)}
ALLOWED_SEXO: Set[str] = {"F", "I", "M"}
ALLOWED_ZONA: Set[str] = {"01", "02"}
ALLOWED_INCAPACIDAD: Set[str] = {"SI", "NO"}

ALLOWED_MODALIDAD: Set[str] = {"01", "02", "03", "04", "06", "07", "08", "09"}
ALLOWED_GRUPO_SERVICIOS: Set[str] = {"01", "02", "03", "04", "05"}
ALLOWED_TIPO_DIAG_PRINCIPAL: Set[str] = {"01", "02", "03"}
ALLOWED_CONCEPTO_RECAUDO: Set[str] = {"01", "02", "03", "05"}
ALLOWED_FINALIDAD: Set[str] = {f"{i:02d}" for i in range(11, 45)}
ALLOWED_CAUSA: Set[str] = {f"{i:02d}" for i in range(21, 49)}


# ============================================================
# UTILIDADES
# ============================================================
def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s


def to_clean_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def resolve_source_column(df_columns: List[str], wanted_name: str) -> Optional[str]:
    if wanted_name in df_columns:
        return wanted_name

    norm_to_real = {normalize_text(c): c for c in df_columns}
    wanted_norm = normalize_text(wanted_name)

    if wanted_norm in norm_to_real:
        return norm_to_real[wanted_norm]

    match = get_close_matches(wanted_norm, list(norm_to_real.keys()), n=1, cutoff=0.75)
    if match:
        return norm_to_real[match[0]]

    return None


def get_target_headers(ws, header_row: int) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val is not None and str(val).strip():
            headers[normalize_text(str(val).strip())] = col
    return headers


def find_target_col(headers_map: Dict[str, int], wanted_name: str) -> Optional[int]:
    return headers_map.get(normalize_text(wanted_name))


def extract_leading_token(value: Any) -> str:
    s = to_clean_text(value)
    if not s:
        return ""
    m = re.match(r"^\s*([A-Za-z0-9]+)", s)
    return m.group(1) if m else ""


def extract_leading_digits(value: Any) -> str:
    s = to_clean_text(value)
    if not s:
        return ""
    m = re.match(r"^\s*(\d+)", s)
    return m.group(1) if m else ""


def normalize_tipo_documento(value: Any) -> str:
    token = extract_leading_token(value).upper()
    return token if token in ALLOWED_TIPO_DOC else ""


def normalize_tipo_usuario(value: Any) -> str:
    digits = extract_leading_digits(value)
    if not digits:
        return ""
    code = digits.zfill(2)
    return code if code in ALLOWED_TIPO_USUARIO else ""


def normalize_sexo(value: Any) -> str:
    s = normalize_text(value)
    map_exact = {
        "f": "F", "femenino": "F", "mujer": "F",
        "m": "M", "masculino": "M", "hombre": "M",
        "i": "I", "indeterminado": "I", "intersexual": "I",
    }
    if s in map_exact:
        return map_exact[s]
    if s.startswith("f"):
        return "F"
    if s.startswith("m"):
        return "M"
    if s.startswith("i"):
        return "I"
    return ""


def normalize_incapacidad(value: Any) -> str:
    s = normalize_text(value)
    if s in {"si", "s", "1", "true", "verdadero"}:
        return "SI"
    if s in {"no", "n", "0", "false", "falso"}:
        return "NO"
    return ""


def normalize_country_or_municipio_code(value: Any) -> str:
    return extract_leading_digits(value)


def normalize_zona(value: Any) -> str:
    digits = extract_leading_digits(value)
    if not digits:
        return ""
    code = digits.zfill(2)
    return code if code in ALLOWED_ZONA else ""


def normalize_fecha_yyyy_mm_dd(value: Any) -> str:
    if pd.isna(value):
        return ""
    dt = pd.to_datetime([value], errors="coerce")[0]
    if pd.isna(dt):
        return ""
    return dt.strftime("%Y-%m-%d")


def normalize_fecha_yyyy_mm_dd_hh_mm(value: Any) -> str:
    if pd.isna(value):
        return ""
    dt = pd.to_datetime([value], errors="coerce")[0]
    if pd.isna(dt):
        return ""
    return dt.strftime("%Y-%m-%d %H:%M")


def normalize_two_digit_code(value: Any, allowed: Set[str]) -> str:
    digits = extract_leading_digits(value)
    if not digits:
        return ""
    code = digits.zfill(2)
    return code if code in allowed else ""


def normalize_cod_servicio(value: Any) -> str:
    return extract_leading_digits(value)


def normalize_amount(value: Any, default_value: str = "0") -> str:
    if pd.isna(value):
        return default_value
    raw = str(value).strip()
    if not raw:
        return default_value

    n = pd.to_numeric(raw.replace(",", ""), errors="coerce")
    if pd.isna(n):
        digits = "".join(re.findall(r"\d", raw))
        return digits if digits else default_value

    return str(int(round(float(n))))


# ============================================================
# VALIDACIONES
# ============================================================
def validate_transaccion(data: Dict[str, Any]) -> None:
    ndio = to_clean_text(data.get("numDocumentoIdObligado"))
    if not re.fullmatch(r"\d{4,12}", ndio):
        raise ValueError(f"numDocumentoIdObligado invalido: {ndio}")

    nf = to_clean_text(data.get("numFactura"))
    if not (1 <= len(nf) <= 20):
        raise ValueError(f"numFactura invalido (1-20): {nf}")

    tn = to_clean_text(data.get("tipoNota")).upper()
    if tn not in ALLOWED_TIPO_NOTA:
        raise ValueError(f"tipoNota invalido: {tn}")

    nn = to_clean_text(data.get("numNota"))
    if nn != "" and not (1 <= len(nn) <= 20):
        raise ValueError(f"numNota invalido (vacio o 1-20): {nn}")


def validate_usuario_row(data: Dict[str, Any]) -> None:
    tdi = to_clean_text(data.get("tipoDocumentoIdentificacion")).upper()
    if tdi not in ALLOWED_TIPO_DOC:
        raise ValueError(f"tipoDocumentoIdentificacion invalido: {tdi}")

    ndi = to_clean_text(data.get("numDocumentoIdentificacion"))
    if not (4 <= len(ndi) <= 20):
        raise ValueError(f"numDocumentoIdentificacion invalido (4-20): {ndi}")

    ndio = to_clean_text(data.get("num_DocumentoIdObligado"))
    if not re.fullmatch(r"\d{4,12}", ndio):
        raise ValueError(f"num_DocumentoIdObligado invalido: {ndio}")

    tu = to_clean_text(data.get("tipoUsuario"))
    if tu not in ALLOWED_TIPO_USUARIO:
        raise ValueError(f"tipoUsuario invalido: {tu}")

    fn = to_clean_text(data.get("fechaNacimiento"))
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", fn):
        raise ValueError(f"fechaNacimiento invalida (yyyy-mm-dd): {fn}")

    sx = to_clean_text(data.get("codSexo")).upper()
    if sx not in ALLOWED_SEXO:
        raise ValueError(f"codSexo invalido: {sx}")

    cpr = to_clean_text(data.get("codPaisResidencia"))
    if not re.fullmatch(r"\d{1,3}", cpr):
        raise ValueError(f"codPaisResidencia invalido: {cpr}")

    cmr = to_clean_text(data.get("codMunicipioResidencia"))
    if not re.fullmatch(r"\d{5}", cmr):
        raise ValueError(f"codMunicipioResidencia invalido: {cmr}")

    ztr = to_clean_text(data.get("codZonaTerritorialResidencia"))
    if ztr not in ALLOWED_ZONA:
        raise ValueError(f"codZonaTerritorialResidencia invalido: {ztr}")

    inc = to_clean_text(data.get("incapacidad")).upper()
    if inc not in ALLOWED_INCAPACIDAD:
        raise ValueError(f"incapacidad invalida: {inc}")

    cpo = to_clean_text(data.get("codPaisOrigen"))
    if not re.fullmatch(r"\d{1,3}", cpo):
        raise ValueError(f"codPaisOrigen invalido: {cpo}")

    cons = to_clean_text(data.get("consecutivo"))
    if not re.fullmatch(r"\d{1,7}", cons):
        raise ValueError(f"consecutivo invalido (1-7 digitos): {cons}")


def validate_consulta_row(data: Dict[str, Any]) -> None:
    ndio = to_clean_text(data.get("num_DocumentoIdObligado"))
    if not re.fullmatch(r"\d{4,12}", ndio):
        raise ValueError(f"num_DocumentoIdObligado invalido: {ndio}")

    cu = to_clean_text(data.get("consecutivoUsuario"))
    if not (4 <= len(cu) <= 20):
        raise ValueError(f"consecutivoUsuario invalido (4-20): {cu}")

    cp = to_clean_text(data.get("codPrestador"))
    if len(cp) != 12:
        raise ValueError(f"codPrestador invalido (12 chars): {cp}")

    fia = to_clean_text(data.get("fechaInicioAtencion"))
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2} \d{2}:\d{2}", fia):
        raise ValueError(f"fechaInicioAtencion invalida (yyyy-mm-dd HH:MM): {fia}")

    na = to_clean_text(data.get("numAutorizacion"))
    if na != "" and not (1 <= len(na) <= 30):
        raise ValueError(f"numAutorizacion invalido (vacio o 1-30): {na}")

    cc = to_clean_text(data.get("codConsulta"))
    if len(cc) != 6:
        raise ValueError(f"codConsulta invalido (6 chars): {cc}")

    mg = to_clean_text(data.get("modalidadGrupoServicioTecSal"))
    if mg not in ALLOWED_MODALIDAD:
        raise ValueError(f"modalidadGrupoServicioTecSal invalido: {mg}")

    gs = to_clean_text(data.get("grupoServicios"))
    if gs not in ALLOWED_GRUPO_SERVICIOS:
        raise ValueError(f"grupoServicios invalido: {gs}")

    cs = to_clean_text(data.get("codServicio"))
    if not re.fullmatch(r"\d{2,4}", cs):
        raise ValueError(f"codServicio invalido (2-4 digitos): {cs}")

    ft = to_clean_text(data.get("finalidadTecnologiaSalud"))
    if ft not in ALLOWED_FINALIDAD:
        raise ValueError(f"finalidadTecnologiaSalud invalido: {ft}")

    cm = to_clean_text(data.get("causaMotivoAtencion"))
    if cm not in ALLOWED_CAUSA:
        raise ValueError(f"causaMotivoAtencion invalido: {cm}")

    cdp = to_clean_text(data.get("codDiagnosticoPrincipal"))
    if not (4 <= len(cdp) <= 25):
        raise ValueError(f"codDiagnosticoPrincipal invalido (4-25): {cdp}")

    for rel in ["codDiagnosticoRelacionado1", "codDiagnosticoRelacionado2", "codDiagnosticoRelacionado3"]:
        val = to_clean_text(data.get(rel))
        if val != "" and not (4 <= len(val) <= 25):
            raise ValueError(f"{rel} invalido (vacio o 4-25): {val}")

    tdp = to_clean_text(data.get("tipoDiagnosticoPrincipal"))
    if tdp not in ALLOWED_TIPO_DIAG_PRINCIPAL:
        raise ValueError(f"tipoDiagnosticoPrincipal invalido: {tdp}")

    tdi = to_clean_text(data.get("tipoDocumentoIdentificacion")).upper()
    if tdi not in ALLOWED_TIPO_DOC:
        raise ValueError(f"tipoDocumentoIdentificacion invalido: {tdi}")

    ndi = to_clean_text(data.get("numDocumentoIdentificacion"))
    if not (4 <= len(ndi) <= 20):
        raise ValueError(f"numDocumentoIdentificacion invalido (4-20): {ndi}")

    vs = to_clean_text(data.get("vrServicio"))
    if not re.fullmatch(r"\d{1,10}", vs):
        raise ValueError(f"vrServicio invalido (1-10 digitos): {vs}")

    cr = to_clean_text(data.get("conceptoRecaudo"))
    if cr not in ALLOWED_CONCEPTO_RECAUDO:
        raise ValueError(f"conceptoRecaudo invalido: {cr}")

    vpm = to_clean_text(data.get("valorPagoModerador"))
    if not re.fullmatch(r"\d{1,10}", vpm):
        raise ValueError(f"valorPagoModerador invalido (1-10 digitos): {vpm}")

    nfev = to_clean_text(data.get("numFEVPagoModerador"))
    if nfev != "" and len(nfev) > 20:
        raise ValueError(f"numFEVPagoModerador invalido (max 20): {nfev}")

    cons = to_clean_text(data.get("consecutivo"))
    if not re.fullmatch(r"\d{1,7}", cons):
        raise ValueError(f"consecutivo invalido (1-7 digitos): {cons}")


class ValidadorRipsService:
    def __init__(
            self,
            config: Dict[str, Any] | None = None,
            on_log: Callable[[str], None] | None = None,
            on_progress: Callable[[int, str], None] | None = None,
    ):
        self.config = config or dict(CONFIG)
        self.on_log = on_log
        self.on_progress = on_progress

        self.transaccion = ProcesadorHojaTransaccion(
            get_target_headers=get_target_headers,
            find_target_col=find_target_col,
            resolve_source_column=resolve_source_column,
            to_clean_text=to_clean_text,
            validate_transaccion=validate_transaccion,
        )
        self.usuarios = ProcesadorHojaUsuarios(
            get_target_headers=get_target_headers,
            find_target_col=find_target_col,
            resolve_source_column=resolve_source_column,
            to_clean_text=to_clean_text,
            normalize_tipo_documento=normalize_tipo_documento,
            normalize_tipo_usuario=normalize_tipo_usuario,
            normalize_fecha_yyyy_mm_dd=normalize_fecha_yyyy_mm_dd,
            normalize_sexo=normalize_sexo,
            normalize_country_or_municipio_code=normalize_country_or_municipio_code,
            normalize_zona=normalize_zona,
            normalize_incapacidad=normalize_incapacidad,
            validate_usuario_row=validate_usuario_row,
        )
        self.consultas = ProcesadorHojaConsultas(
            get_target_headers=get_target_headers,
            find_target_col=find_target_col,
            resolve_source_column=resolve_source_column,
            to_clean_text=to_clean_text,
            normalize_fecha_yyyy_mm_dd_hh_mm=normalize_fecha_yyyy_mm_dd_hh_mm,
            normalize_two_digit_code=normalize_two_digit_code,
            normalize_cod_servicio=normalize_cod_servicio,
            normalize_tipo_documento=normalize_tipo_documento,
            normalize_amount=normalize_amount,
            validate_consulta_row=validate_consulta_row,
            allowed_modalidad=ALLOWED_MODALIDAD,
            allowed_grupo_servicios=ALLOWED_GRUPO_SERVICIOS,
            allowed_finalidad=ALLOWED_FINALIDAD,
            allowed_causa=ALLOWED_CAUSA,
            allowed_tipo_diag_principal=ALLOWED_TIPO_DIAG_PRINCIPAL,
            allowed_concepto_recaudo=ALLOWED_CONCEPTO_RECAUDO,
        )

    def _log(self, mensaje: str) -> None:
        if self.on_log:
            self.on_log(mensaje)
        else:
            print(mensaje)

    def _progress(self, porcentaje: int, estado: str) -> None:
        if self.on_progress:
            self.on_progress(porcentaje, estado)

    def ejecutar(self) -> Path:
        source_file = Path(str(self.config["source_file"]))
        template_file = Path(str(self.config["template_file"]))
        output_file = Path(str(self.config["output_file"]))

        self._progress(5, "Leyendo archivo origen")
        df_source = pd.read_excel(source_file, sheet_name=int(self.config["source_sheet"]))

        self._progress(20, "Abriendo plantilla")
        wb = load_workbook(template_file, keep_vba=True)

        self._progress(40, "Procesando hoja transaccion")
        ndio = self.transaccion.ejecutar(df_source, wb, self.config)

        self._progress(65, "Procesando hoja usuarios")
        usuarios_ids = self.usuarios.ejecutar(df_source, wb, self.config, str(ndio))

        self._progress(85, "Procesando hoja consultas")
        self.consultas.ejecutar(df_source, wb, self.config, str(ndio), usuarios_ids)

        self._progress(95, "Guardando archivo")
        wb.save(output_file)

        self._progress(100, "Proceso finalizado")
        self._log(f"Archivo generado: {output_file}")
        return output_file


def main() -> None:
    output = ValidadorRipsService(CONFIG).ejecutar()
    print(f"Archivo generado: {output}")


if __name__ == "__main__":
    main()
